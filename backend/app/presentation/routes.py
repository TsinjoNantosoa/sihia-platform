from dataclasses import asdict
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import Response, StreamingResponse

from app.application.schemas import (
    AppointmentCreate,
    AppointmentReschedule,
    AppointmentStatusUpdate,
    ReminderSendRequest,
    ForgotPasswordRequest,
    LoginRequest,
    LoginResponse,
    LogoutRequest,
    MedicalVisitCreate,
    NotificationMarkReadRequest,
    NotificationPrefsUpdate,
    PatientCreate,
    DoctorUpdate,
    PatientUpdate,
    RefreshTokenRequest,
    ResetPasswordRequest,
    UserCreate,
    UserUpdate,
    VerifyResetCodeRequest,
)
from app.presentation.deps import (
    analytics_service,
    appointments_service,
    waiting_room_service,
    auth_service,
    doctors_service,
    medical_history_service,
    patients_service,
    ml_service,
    noshow_service,
    notification_service,
    patient_summary_service,
    patient_document_service,
    search_service,
    pipeline_service,
    rbac_service,
    reminder_service,
    require_auth,
    require_permission,
)
from app.infrastructure.audit_log import export_audit_jsonl, read_audit_records
from app.presentation.audit import log_admin_action
from app.presentation.rate_limit import (
    check_login_allowed,
    check_password_reset_allowed,
    register_login_failure,
    register_password_reset_attempt,
    reset_login_limiter,
)


def _doctor_payload(d) -> dict:
    return {
        "id": d.id,
        "firstName": d.first_name,
        "lastName": d.last_name,
        "specialty": d.specialty,
        "phone": d.phone,
        "email": d.email,
        "availability": d.availability,
        "patientsCount": d.patients_count,
        "weeklyAppointments": d.weekly_appointments,
        "satisfaction": d.satisfaction,
        "schedule": d.schedule,
    }


def _patient_payload(p) -> dict:
    return {
        "id": p.id,
        "recordNumber": p.record_number,
        "firstName": p.first_name,
        "lastName": p.last_name,
        "dob": p.dob,
        "gender": p.gender,
        "phone": p.phone,
        "email": p.email,
        "address": p.address,
        "bloodType": p.blood_type,
        "allergies": p.allergies,
        "insurance": p.insurance,
        "status": p.status,
        "lastVisit": p.last_visit,
        "chronicConditions": p.chronic_conditions,
        "currentTreatments": p.current_treatments,
        "emergencyContact": p.emergency_contact,
    }

api_router = APIRouter(prefix="/api")


@api_router.post("/auth/login", response_model=LoginResponse)
def login(payload: LoginRequest, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    limiter_key = f"{client_ip}:{payload.email.lower()}"
    retry_after = check_login_allowed(limiter_key)
    if retry_after is not None:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail={
                "code": "TOO_MANY_ATTEMPTS",
                "message": "Trop de tentatives de connexion. Réessayez plus tard.",
                "retryAfterSeconds": retry_after,
            },
        )

    try:
        access_token, refresh_token = auth_service.login(payload.email, payload.password)
    except HTTPException as exc:
        if exc.status_code == status.HTTP_401_UNAUTHORIZED:
            register_login_failure(limiter_key)
        raise

    reset_login_limiter(limiter_key)
    return LoginResponse(access_token=access_token, refresh_token=refresh_token)


@api_router.post("/auth/refresh", response_model=LoginResponse)
def refresh_token(payload: RefreshTokenRequest):
    access_token, refresh_token_value = auth_service.refresh(payload.refresh_token)
    return LoginResponse(access_token=access_token, refresh_token=refresh_token_value)


@api_router.post("/auth/logout")
def logout(payload: LogoutRequest):
    auth_service.logout(payload.refresh_token)
    return {"success": True}


@api_router.post("/auth/logout-all")
def logout_all(request: Request, claims: dict = Depends(require_auth)):
    auth_service.logout_all(claims.get("sub"))
    log_admin_action(
        request,
        action="auth.logout_all",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
    )
    return {"success": True}


@api_router.post("/auth/forgot-password")
def forgot_password(payload: ForgotPasswordRequest, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    limiter_key = f"reset:{client_ip}:{payload.email.lower()}"
    retry_after = check_password_reset_allowed(limiter_key)
    if retry_after is not None:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail={
                "code": "TOO_MANY_ATTEMPTS",
                "message": "Trop de tentatives. Réessayez plus tard.",
                "retryAfterSeconds": retry_after,
            },
        )
    register_password_reset_attempt(limiter_key)
    auth_service.request_password_reset(str(payload.email))
    return {
        "status": "ok",
        "message": "Si ce compte existe, un code de vérification a été envoyé.",
    }


@api_router.post("/auth/verify-reset-code")
def verify_reset_code(payload: VerifyResetCodeRequest, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    limiter_key = f"reset:{client_ip}:{payload.email.lower()}"
    retry_after = check_password_reset_allowed(limiter_key)
    if retry_after is not None:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail={
                "code": "TOO_MANY_ATTEMPTS",
                "message": "Trop de tentatives. Réessayez plus tard.",
                "retryAfterSeconds": retry_after,
            },
        )
    register_password_reset_attempt(limiter_key)
    auth_service.verify_reset_code(str(payload.email), payload.code)
    return {"status": "ok"}


@api_router.post("/auth/reset-password")
def reset_password(payload: ResetPasswordRequest, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    limiter_key = f"reset:{client_ip}:{payload.email.lower()}"
    retry_after = check_password_reset_allowed(limiter_key)
    if retry_after is not None:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail={
                "code": "TOO_MANY_ATTEMPTS",
                "message": "Trop de tentatives. Réessayez plus tard.",
                "retryAfterSeconds": retry_after,
            },
        )
    register_password_reset_attempt(limiter_key)
    auth_service.reset_password(str(payload.email), payload.code, payload.new_password)
    return {"status": "ok"}


@api_router.get("/patients")
def list_patients(
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    _claims: dict = Depends(require_permission("patients:read")),
):
    items = patients_service.list(search=search, status_filter=status)
    return [_patient_payload(p) for p in items]


@api_router.get("/patients/{patient_id}")
def get_patient(patient_id: str, _claims: dict = Depends(require_permission("patients:read"))):
    return _patient_payload(patients_service.get(patient_id))


@api_router.post("/patients")
def create_patient(
    request: Request,
    payload: PatientCreate,
    claims: dict = Depends(require_permission("patients:create")),
):
    patient = patients_service.create(payload)
    log_admin_action(
        request,
        action="patient.create",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient.id,
    )
    return _patient_payload(patient)


@api_router.patch("/patients/{patient_id}")
def update_patient(
    request: Request,
    patient_id: str,
    payload: PatientUpdate,
    claims: dict = Depends(require_permission("patients:update")),
):
    patient = patients_service.update(patient_id, payload)
    log_admin_action(
        request,
        action="patient.update",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient_id,
    )
    return _patient_payload(patient)


@api_router.delete("/patients/{patient_id}", status_code=status.HTTP_200_OK)
def delete_patient(
    request: Request,
    patient_id: str,
    claims: dict = Depends(require_permission("patients:delete")),
):
    patients_service.delete(patient_id)
    log_admin_action(
        request,
        action="patient.archive",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient_id,
    )
    return {"success": True}


@api_router.get("/patients/{patient_id}/history")
def get_patient_history(patient_id: str, _claims: dict = Depends(require_permission("patients:read"))):
    visits = medical_history_service.list(patient_id)
    return [
        {
            "id": v.id,
            "date": v.date,
            "reason": v.reason,
            "doctorName": v.doctor_name,
            "specialty": v.specialty,
            "diagnosis": v.diagnosis,
            "treatment": v.treatment,
            "notes": v.notes,
        }
        for v in visits
    ]


@api_router.post("/patients/{patient_id}/ai-summary")
def patient_ai_summary(
    request: Request,
    patient_id: str,
    lang: str = Query("fr", pattern="^(fr|en|ar)$"),
    claims: dict = Depends(require_permission("patients:read")),
):
    """Résumé IA du dossier (~5 lignes) — aide à la décision, pas un diagnostic."""
    patient = patients_service.get(patient_id)
    visits = medical_history_service.list(patient_id)
    result = patient_summary_service.summarize(patient, visits, lang=lang)
    log_admin_action(
        request,
        action="patient.ai_summary.generate",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient_id,
        extra={"lang": lang},
    )
    return result


@api_router.get("/patients/{patient_id}/documents")
def list_patient_documents(patient_id: str, _claims: dict = Depends(require_permission("patients:read"))):
    return patient_document_service.list(patient_id)


@api_router.post("/patients/{patient_id}/documents", status_code=status.HTTP_201_CREATED)
async def upload_patient_document(
    request: Request,
    patient_id: str,
    file: UploadFile = File(...),
    category: str = Form("other"),
    notes: str | None = Form(None),
    claims: dict = Depends(require_permission("patients:update")),
):
    data = await file.read()
    doc = patient_document_service.upload(
        patient_id,
        filename=file.filename or "document",
        content_type=file.content_type or "application/octet-stream",
        data=data,
        category=category,
        notes=notes,
        uploaded_by=str(claims.get("id") or claims.get("sub") or ""),
    )
    log_admin_action(
        request,
        action="patient.document.upload",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient_id,
        extra={
            "documentId": doc.get("id"),
            "filename": doc.get("filename"),
            "category": category,
        },
    )
    return doc


@api_router.get("/patients/{patient_id}/documents/{document_id}/download")
def download_patient_document(
    request: Request,
    patient_id: str,
    document_id: str,
    claims: dict = Depends(require_permission("patients:read")),
):
    doc, payload = patient_document_service.get_file(patient_id, document_id)
    log_admin_action(
        request,
        action="patient.document.download",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient_id,
        extra={"documentId": document_id, "filename": doc.filename},
    )
    return Response(
        content=payload,
        media_type=doc.content_type,
        headers={"Content-Disposition": f'inline; filename="{doc.filename}"'},
    )


@api_router.delete("/patients/{patient_id}/documents/{document_id}", status_code=status.HTTP_200_OK)
def delete_patient_document(
    request: Request,
    patient_id: str,
    document_id: str,
    claims: dict = Depends(require_permission("patients:update")),
):
    patient_document_service.delete(patient_id, document_id)
    log_admin_action(
        request,
        action="patient.document.delete",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient_id,
        extra={"documentId": document_id},
    )
    return {"ok": True}


@api_router.post("/patients/{patient_id}/history")
def add_patient_visit(
    request: Request,
    patient_id: str,
    payload: MedicalVisitCreate,
    claims: dict = Depends(require_permission("patients:update")),
):
    v = medical_history_service.add(patient_id, payload)
    log_admin_action(
        request,
        action="medical_visit.create",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=patient_id,
        extra={"visitId": v.id},
    )
    return {
        "id": v.id,
        "date": v.date,
        "reason": v.reason,
        "doctorName": v.doctor_name,
        "specialty": v.specialty,
        "diagnosis": v.diagnosis,
        "treatment": v.treatment,
        "notes": v.notes,
    }


@api_router.get("/doctors")
def list_doctors(_claims: dict = Depends(require_permission("doctors:read"))):
    return [_doctor_payload(d) for d in doctors_service.list()]


@api_router.get("/doctors/{doctor_id}")
def get_doctor(doctor_id: str, _claims: dict = Depends(require_permission("doctors:read"))):
    return _doctor_payload(doctors_service.get(doctor_id))


@api_router.patch("/doctors/{doctor_id}")
def update_doctor(
    doctor_id: str,
    payload: DoctorUpdate,
    _claims: dict = Depends(require_permission("doctors:update")),
):
    return _doctor_payload(doctors_service.update(doctor_id, payload))


def _appointment_payload(a, reminder_summary: dict | None = None) -> dict:
    payload = {
        "id": a.id,
        "patientId": a.patient_id,
        "patientName": a.patient_name,
        "doctorId": a.doctor_id,
        "doctorName": a.doctor_name,
        "date": a.date,
        "durationMin": a.duration_min,
        "reason": a.reason,
        "status": a.status,
    }
    if reminder_summary is not None:
        payload["reminderSummary"] = reminder_summary
    return payload


@api_router.get("/appointments")
def list_appointments(_claims: dict = Depends(require_permission("appointments:read"))):
    appointments = appointments_service.list()
    summaries = reminder_service.summaries_for([a.id for a in appointments])
    return [_appointment_payload(a, summaries.get(a.id)) for a in appointments]


@api_router.post("/appointments")
def create_appointment(
    request: Request,
    payload: AppointmentCreate,
    claims: dict = Depends(require_permission("appointments:create")),
):
    a = appointments_service.create(payload)
    log_admin_action(
        request,
        action="appointment.create",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=a.id,
        extra={"patientId": a.patient_id, "doctorId": a.doctor_id},
    )
    return _appointment_payload(a)


@api_router.post("/appointments/{appointment_id}/cancel")
def cancel_appointment(
    request: Request,
    appointment_id: str,
    claims: dict = Depends(require_permission("appointments:update")),
):
    a = appointments_service.cancel(appointment_id)
    log_admin_action(
        request,
        action="appointment.cancel",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=appointment_id,
    )
    return asdict(a)


@api_router.patch("/appointments/{appointment_id}/status")
def update_appointment_status(
    request: Request,
    appointment_id: str,
    payload: AppointmentStatusUpdate,
    claims: dict = Depends(require_permission("appointments:update")),
):
    previous = appointments_service.appointments.get(appointment_id)
    updated = appointments_service.transition_status(appointment_id, payload.status)
    log_admin_action(
        request,
        action="appointments.status.update",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=appointment_id,
        extra={"from": previous.status if previous else None, "to": updated.status},
    )
    return _appointment_payload(updated)


@api_router.patch("/appointments/{appointment_id}/schedule")
def reschedule_appointment(
    request: Request,
    appointment_id: str,
    payload: AppointmentReschedule,
    claims: dict = Depends(require_permission("appointments:update")),
):
    previous = appointments_service.appointments.get(appointment_id)
    doctor = doctors_service.get(payload.doctorId)
    doctor_name = f"Dr. {doctor.first_name} {doctor.last_name}"
    updated = appointments_service.reschedule(
        appointment_id,
        doctor_id=doctor.id,
        doctor_name=doctor_name,
        date=payload.date,
    )
    log_admin_action(
        request,
        action="appointments.schedule.update",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=appointment_id,
        extra={
            "fromDoctorId": previous.doctor_id if previous else None,
            "toDoctorId": updated.doctor_id,
            "fromDate": previous.date if previous else None,
            "toDate": updated.date,
        },
    )
    return _appointment_payload(updated)


@api_router.get("/appointments/{appointment_id}/reminders")
def list_appointment_reminders(
    appointment_id: str,
    _claims: dict = Depends(require_permission("appointments:read")),
):
    reminders = reminder_service.list_for_appointment(appointment_id)
    return {
        "items": [
            {
                "id": r.id,
                "channel": r.channel,
                "kind": r.kind,
                "status": r.status,
                "recipient": r.recipient,
                "sentAt": r.sent_at,
                "error": r.error,
            }
            for r in reminders
        ]
    }


@api_router.post("/appointments/{appointment_id}/remind")
def send_appointment_reminder(
    request: Request,
    appointment_id: str,
    payload: ReminderSendRequest,
    claims: dict = Depends(require_permission("appointments:update")),
):
    result = reminder_service.send_manual(appointment_id, payload.channels)
    log_admin_action(
        request,
        action="appointments.remind",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=appointment_id,
        extra={"channels": payload.channels, "results": result.get("results")},
    )
    return result


@api_router.post("/admin/reminders/run")
def run_appointment_reminders_batch(
    request: Request,
    claims: dict = Depends(require_permission("appointments:update")),
):
    result = reminder_service.run_auto_batch()
    log_admin_action(
        request,
        action="appointments.reminders.batch",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        extra=result,
    )
    return result


@api_router.get("/admin/reminders/status")
def reminders_status(_claims: dict = Depends(require_permission("appointments:update"))):
    from app.infrastructure.notification_channels import reminder_channels_status

    return reminder_channels_status()


@api_router.get("/admin/pipeline/status")
def pipeline_status(_claims: dict = Depends(require_permission("analytics:read"))):
    return pipeline_service.status()


@api_router.post("/admin/pipeline/run/{dag_id}")
def run_pipeline_dag(
    request: Request,
    dag_id: str,
    claims: dict = Depends(require_permission("appointments:update")),
):
    try:
        result = pipeline_service.run(dag_id)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(exc)) from exc
    log_admin_action(
        request,
        action="pipeline.run",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        extra={"dagId": dag_id, "result": result},
    )
    return result


@api_router.get("/analytics/kpis")
def kpis(_claims: dict = Depends(require_permission("analytics:read"))):
    return analytics_service.kpis()


@api_router.get("/analytics/revenue")
def monthly_revenue(period: str = Query(default="6m"), _claims: dict = Depends(require_permission("analytics:read"))):
    return analytics_service.monthly_revenue(period)


@api_router.get("/analytics/admissions-dept")
def admissions_by_dept(_claims: dict = Depends(require_permission("analytics:read"))):
    return analytics_service.admissions_by_dept()


@api_router.get("/analytics/satisfaction")
def satisfaction(_claims: dict = Depends(require_permission("analytics:read"))):
    return analytics_service.satisfaction()


@api_router.get("/analytics/export/excel")
def export_excel(period: str = Query(default="6m"), _claims: dict = Depends(require_permission("analytics:read"))):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analytics Export"

    ws.append(["Rapport Analytique - SIH IA"])
    ws.append(["Période", period])
    ws.append([])

    ws.append(["Indicateurs Clés"])
    ws.append(["Indicateur", "Valeur"])
    kpi_data = analytics_service.kpis()
    ws.append(["Patients Aujourd'hui", kpi_data["patientsToday"]])
    ws.append(["Taux d'occupation", f"{kpi_data['occupancy']}%"])
    ws.append(["Total Rendez-vous", kpi_data["appointments"]])
    ws.append([])

    ws.append(["Revenus mensuels"])
    ws.append(["Mois", "Revenu (€)"])
    revenue_data = analytics_service.monthly_revenue(period)
    for r in revenue_data:
        ws.append([r["label"], r["value"]])

    ws2 = wb.create_sheet(title="Patients")
    ws2.append(["ID Dossier", "Nom", "Prenom", "Telephone", "Statut", "Derniere Visite"])
    for p in patients_service.list(None, None):
        ws2.append([p.record_number, p.last_name, p.first_name, p.phone, p.status, p.last_visit])

    ws3 = wb.create_sheet(title="Rendez-vous")
    ws3.append(["ID", "Patient", "Medecin", "Date", "Motif", "Statut"])
    for a in appointments_service.list():
        ws3.append([a.id, a.patient_name, a.doctor_name, a.date, a.reason, a.status])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=analytics_{period}.xlsx"}
    )


@api_router.get("/analytics/export/pdf")
def export_pdf(period: str = Query(default="6m"), _claims: dict = Depends(require_permission("analytics:read"))):
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("helvetica", size=16)
    pdf.cell(0, 10, text="Rapport Analytique - SIH IA", ln=True, align="C")
    
    pdf.set_font("helvetica", size=12)
    pdf.ln(10)
    pdf.cell(0, 10, text=f"Periode : {period}", ln=True)
    pdf.ln(5)

    pdf.set_font("helvetica", style="B", size=14)
    pdf.cell(0, 10, text="Indicateurs Cles", ln=True)
    pdf.set_font("helvetica", size=12)
    kpi_data = analytics_service.kpis()
    pdf.cell(0, 10, text=f"Patients Aujourd'hui : {kpi_data['patientsToday']}", ln=True)
    pdf.cell(0, 10, text=f"Taux d'occupation : {kpi_data['occupancy']}%", ln=True)
    pdf.cell(0, 10, text=f"Total Rendez-vous : {kpi_data['appointments']}", ln=True)
    pdf.ln(5)

    pdf.set_font("helvetica", style="B", size=14)
    pdf.cell(0, 10, text="Revenus mensuels", ln=True)
    pdf.set_font("helvetica", size=12)
    revenue_data = analytics_service.monthly_revenue(period)
    for r in revenue_data:
        pdf.cell(0, 10, text=f"{r['label']} : {r['value']} eur", ln=True)

    pdf.add_page()
    pdf.set_font("helvetica", style="B", size=14)
    pdf.cell(0, 10, text="Registre des Patients (Extrait)", ln=True)
    pdf.set_font("helvetica", size=10)
    pdf.cell(40, 10, text="Dossier", border=1)
    pdf.cell(60, 10, text="Nom Prenom", border=1)
    pdf.cell(40, 10, text="Telephone", border=1)
    pdf.cell(40, 10, text="Statut", border=1, ln=True)
    
    # Export up to 100 recent patients for PDF to avoid blowing up memory quickly
    all_patients = patients_service.list(None, None)[:100]
    for p in all_patients:
        pdf.cell(40, 10, text=str(p.record_number), border=1)
        pdf.cell(60, 10, text=f"{p.last_name[:15]} {p.first_name[:15]}", border=1)
        pdf.cell(40, 10, text=str(p.phone), border=1)
        pdf.cell(40, 10, text=str(p.status), border=1, ln=True)

    out = BytesIO(pdf.output())
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=analytics_{period}.pdf"}
    )


@api_router.get("/ml/predict-7d")
def predict(_claims: dict = Depends(require_permission("ml:read"))):
    return ml_service.predict_7d()


@api_router.get("/ml/predict-30d")
def predict_30d(_claims: dict = Depends(require_permission("ml:read"))):
    return ml_service.predict_30d()


@api_router.get("/ml/metrics")
def ml_metrics(_claims: dict = Depends(require_permission("ml:read"))):
    return ml_service.metrics()


@api_router.get("/ml/noshow-risk")
def noshow_risk(
    horizonDays: int = Query(7, ge=1, le=60),
    minRisk: float = Query(0.0, ge=0.0, le=1.0),
    riskLevel: str | None = Query(None, pattern="^(high|medium|low)$"),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    _claims: dict = Depends(require_permission("ml:read")),
):
    """Liste des RDV à risque d'absence (score heuristique, aide à la décision)."""
    return noshow_service.list_risks(
        horizon_days=horizonDays,
        min_risk=minRisk,
        risk_level=riskLevel,
        limit=limit,
        offset=offset,
    )


@api_router.get("/waiting-room")
def waiting_room(_claims: dict = Depends(require_permission("appointments:read"))):
    return waiting_room_service.snapshot()


@api_router.post("/waiting-room/call-next")
def waiting_room_call_next(
    doctorId: str | None = Query(None),
    _claims: dict = Depends(require_permission("appointments:update")),
):
    return waiting_room_service.call_next(doctor_id=doctorId)


@api_router.get("/alerts")
def alerts(_claims: dict = Depends(require_permission("dashboard:read"))):
    return analytics_service.alerts()


@api_router.get("/search")
def global_search(
    q: str = Query("", max_length=120),
    limit: int = Query(8, ge=1, le=20),
    _claims: dict = Depends(require_permission("dashboard:read")),
):
    """Recherche globale patients / médecins / RDV."""
    return search_service.search(q, limit=limit)


@api_router.get("/notifications")
def list_notifications(
    level: str | None = Query(None, pattern="^(critical|warning|info)$"),
    unreadOnly: bool = Query(False),
    area: str | None = Query(None, max_length=80),
    claims: dict = Depends(require_permission("dashboard:read")),
):
    user_id = str(claims.get("id") or claims.get("sub") or "")
    return notification_service.list_for_user(
        user_id,
        level=level,
        unread_only=unreadOnly,
        area=area,
    )


@api_router.post("/notifications/read")
def mark_notifications_read(
    payload: NotificationMarkReadRequest,
    claims: dict = Depends(require_permission("dashboard:read")),
):
    user_id = str(claims.get("id") or claims.get("sub") or "")
    return notification_service.mark_read(user_id, payload.alertIds)


@api_router.post("/notifications/read-all")
def mark_all_notifications_read(claims: dict = Depends(require_permission("dashboard:read"))):
    user_id = str(claims.get("id") or claims.get("sub") or "")
    return notification_service.mark_all_read(user_id)


@api_router.get("/notifications/prefs")
def get_notification_prefs(claims: dict = Depends(require_permission("settings:read"))):
    user_id = str(claims.get("id") or claims.get("sub") or "")
    return notification_service.get_prefs(user_id)


@api_router.patch("/notifications/prefs")
def update_notification_prefs(
    payload: NotificationPrefsUpdate,
    claims: dict = Depends(require_permission("settings:read")),
):
    user_id = str(claims.get("id") or claims.get("sub") or "")
    return notification_service.update_prefs(user_id, payload.model_dump(exclude_none=True))


@api_router.get("/rbac/users")
def rbac_users(_claims: dict = Depends(require_permission("users:read"))):
    return rbac_service.list_users()


@api_router.post("/rbac/users", status_code=201)
def create_rbac_user(
    request: Request,
    payload: UserCreate,
    claims: dict = Depends(require_permission("users:create")),
):
    created = rbac_service.create_user(payload)
    log_admin_action(
        request,
        action="rbac.user.create",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=created.get("id"),
        extra={"role": created.get("role"), "status": created.get("status")},
    )
    return created


@api_router.patch("/rbac/users/{user_id}")
def update_rbac_user(
    request: Request,
    user_id: str,
    payload: UserUpdate,
    claims: dict = Depends(require_permission("users:update")),
):
    updated = rbac_service.update_user(user_id, payload, claims.get("sub", ""))
    log_admin_action(
        request,
        action="rbac.user.update",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=user_id,
        extra={"role": updated.get("role"), "status": updated.get("status")},
    )
    return updated


@api_router.get("/admin/audit-logs")
def list_audit_logs(
    request: Request,
    limit: int = Query(default=100, ge=1, le=1000),
    claims: dict = Depends(require_permission("users:read")),
):
    items = read_audit_records(limit=limit)
    log_admin_action(
        request,
        action="audit.logs.list",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        extra={"limit": limit, "count": len(items)},
    )
    return {"items": items, "count": len(items)}


@api_router.get("/admin/audit-logs/export")
def export_audit_logs(
    request: Request,
    limit: int = Query(default=5000, ge=1, le=20000),
    claims: dict = Depends(require_permission("users:read")),
):
    log_admin_action(
        request,
        action="audit.logs.export",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        extra={"limit": limit},
    )
    content = export_audit_jsonl(limit=limit)
    stamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"sihia_audit_{stamp}.jsonl"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/x-ndjson",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.delete("/rbac/users/{user_id}", status_code=204)
def delete_rbac_user(
    request: Request,
    user_id: str,
    claims: dict = Depends(require_permission("users:delete")),
):
    rbac_service.delete_user(user_id, claims.get("sub", ""))
    log_admin_action(
        request,
        action="rbac.user.delete",
        actor_id=claims.get("sub"),
        actor_email=claims.get("email"),
        target_id=user_id,
    )
